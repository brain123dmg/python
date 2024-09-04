import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_workbook(filename):
    # Load the workbook and select the active sheet
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # Process rows to correct prices
    for row in range (2, sheet.max_row+1):
        cell = sheet.cell(row,3)
        corrected_price = 0.9 * cell.value
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price
     # Define the data for the bar chart
    values = Reference(sheet,
                        min_row=2,
                        max_row=sheet.max_row,
                        min_col=4,
                        max_col=4)
    # Create the bar chart
    chart = BarChart()
    chart.add_data(values)
    #bar_chart().add_data(values)
    sheet.add_chart(chart,'e2')
    # Save the modified workbook
    wb.save(filename)